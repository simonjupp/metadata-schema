from optparse import OptionParser
import logging
from openpyxl import Workbook
import requests, pprint
from openpyxl.styles import Font
import os
from schema_test_suite import get_json_from_file


# hard coded tab ordering
tab_ordering = ["project", "project.publications", "project.contributors", "donor_organism", "familial_relationship", "specimen_from_organism", "cell_suspension",
                "cell_line", "cell_line.publications", "organoid", "collection_process", "dissociation_process", "enrichment_process", "library_preparation_process",
                "sequencing_process", "purchased_reagents", "protocol", "sequence_file"]

excluded_fields = ["describedBy", "schema_version", "schema_type"]

class SpreadsheetCreator:

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def generateSpreadsheet(self, baseUri, schemas, dependencies, output, local, userFriendly):
        values = {}
        try:
            # for each schema, gather the values for the relevant tab(s)
            for schema in schemas:
                v = self._gatherValues(baseUri, schema, dependencies, local, userFriendly)
                values.update(v)
            # Build the spreadsheet from the retrieved values
            self._buildSpreadsheet(values, output)
        except ValueError as e:
            self.logger.error("Error:" + str(e))
            raise e

    def _gatherValues(self, basepath, schema, dependencies, local, userFriendly):

        if local:
            jsonRaw = get_json_from_file(schema)

        else:
            # get the schema of HTTP
            req = requests.get(basepath+schema)

            # if the schema is successfully retrieved, process it, else return an error message
            if (req.status_code == requests.codes.ok):
                jsonRaw = req.json()
            else:
                self.logger.error(schema + " does not exist")

        entities = {}
        entity_title = jsonRaw["title"]

        properties = jsonRaw["properties"]

        values = []

        for prop in properties:
            # if a property has an array of references (potential 1-to-many relationship), gather the properties for the references and format them to become
            # their own spreadsheet tab
            if ("items" in properties[prop] and "$ref" in properties[prop]["items"] and "ontology" not in properties[prop]["items"]["$ref"]):
                module = properties[prop]["items"]["$ref"]
                if local:
                    el = module.split("/")
                    del el[0:3]
                    del el[-2]

                    module = basepath

                    for e in el:
                        module = module + "/" + e
                    module = module + ".json"
                if module in dependencies:
                    module_values = self._gatherValues(basepath, module, None, local, userFriendly)
                    # add primary entity ID to cross reference with main entity
                    for primary in values:
                        if "id" in primary["header"].lower() or "shortname" in primary["header"]:
                            for key in module_values.keys():
                                t = primary["header"]
                                if "ID" in t:
                                    t = t.replace(" ID", "").lower()
                                    d = "ID for " + t + " this " + key + " relates to"
                                else:
                                    t = t.replace(" shortname", "").lower()
                                    d = "Shortname for " + t + " this " + key + " relates to"
                                module_values[key].append({"header": primary["header"],
                                                           "description": d,
                                                           "example": None})
                            break

                    # special name cases for publication tabs
                    if entity_title == "project" and "publication" in module_values.keys():
                        module_values["project.publications"] = module_values.pop("publication")
                    if entity_title == "cell_line" and "publication" in module_values.keys():
                        module_values["cell_line.publications"] = module_values.pop("publication")

                    entities.update(module_values)
            # if a property does not include a user_friendly tag but includes a reference, fetch the contents of that reference and add them
            # directly to the properties for this sheet
            elif("$ref" in properties[prop] and "ontology" not in properties[prop]["$ref"]):
                module = properties[prop]["$ref"]
                if local:
                    el = module.split("/")
                    del el[0:3]
                    del el[-2]

                    module = basepath

                    for e in el:
                        module = module + "/" + e
                    module = module + ".json"

                    if "_core" in module or module in dependencies:
                        module_values = self._gatherValues(basepath, module, None, local, userFriendly)

                        prefix = ""
                    # if module in dependencies:
                        if userFriendly:
                            if "user_friendly" in properties[prop]:
                                prefix = properties[prop]["user_friendly"] + " - "
                            else:
                                print(prop + " in " + entity_title + " has no user friendly name")
                        else:
                            prefix = prop + "."

                    for key in module_values.keys():
                        for entry in module_values[key]:
                            entry["header"] = prefix + entry["header"]
                        values.extend(module_values[key])

            # if a property has a user_friendly tag, include it as a direct field. This includes ontology module references as these should not be
            # exposed to users
            elif (userFriendly and "user_friendly" in properties[prop]):
                description = None
                example = None
                if "description" in properties[prop]:
                    description = properties[prop]["description"]
                if "example" in properties[prop]:
                    example = properties[prop]["example"]

                values.append({"header": properties[prop]["user_friendly"], "description": description,
                               "example": example})
            elif not userFriendly:
                if prop not in excluded_fields:
                    description = None
                    example = None
                    if "description" in properties[prop]:
                        description = properties[prop]["description"]
                    if "example" in properties[prop]:
                        example = properties[prop]["example"]

                    if(("$ref" in properties[prop] and "ontology" in properties[prop]["$ref"]) or
                            (("items" in properties[prop] and "$ref" in properties[prop]["items"]) and ("ontology" in properties[prop]["items"]["$ref"]))):
                        prop = prop + ".text"

                    values.append({"header": prop, "description": description,
                                   "example": example})

        if "type/biomaterial" in schema:
            if userFriendly:
                values.append(
                    {"header": "Process IDs", "description": "IDs of processes for which this biomaterial is an input",
                                   "example": None})
            else:
                values.append(
                    {"header": "process_ids", "description": "IDs of processes for which this biomaterial is an input",
                     "example": None})
        if "type/process" in schema:
            if userFriendly:
                values.append(
                    {"header": "Protocol IDs", "description": "IDs of protocols which this process implements",
                     "example": None})
            else:
                values.append(
                    {"header": "protocol_ids", "description": "IDs of protocols which this process implements",
                     "example": None})

        if "type/file" in schema:
            if userFriendly:
                values.append(
                    {"header": "Biomaterial ID", "description": "ID of the biomaterial to which this file relates",
                     "example": None})
                values.append(
                    {"header": "Sequencing process ID", "description": "ID of the sequencing process to which this file relates",
                    "example": None})
            else:
                values.append(
                    {"header": "biomaterial_id", "description": "ID of the biomaterial to which this file relates",
                     "example": None})
                values.append(
                    {"header": "process_id",
                     "description": "ID of the sequencing process to which this file relates",
                     "example": None})

        entities[entity_title] = values
        return entities




    def _buildSpreadsheet(self, values, outputLocation):
        wb = Workbook()

        # for each tab entry in the values dictionary, create a new worksheet
        # for tab_name in values.keys():
        for tab_name in tab_ordering:
            if tab_name in values.keys():
                headers = values[tab_name]

                ws = wb.create_sheet(title=tab_name)
                col = 1

                # Optional set of descriptors what each of the 3 top rows contains
                # ws.cell(column=col, row=1, value="Description")
                # ws.cell(column=col, row=2, value="Example").font = Font(italic=True)
                # ws.cell(column=col, row=3, value="Header").font = Font(bold=True)
                # col +=1


                # put each description in row 1, example in row 2 and header in row 3, then increment the column index
                for header in headers:
                    ws.cell(column=col, row=1, value=header["description"])
                    ws.cell(column=col, row=2, value=header["example"]).font = Font(italic = True)
                    ws.cell(column=col, row=3, value=header["header"]).font = Font(bold = True)

                    # set column width to the width of the header or 5 for very short headers (eg DOI)
                    min_width = 5
                    width = len(header["header"])
                    if width < min_width:
                        adjusted_width = min_width
                    else:
                        adjusted_width = width
                    ws.column_dimensions[ws.cell(column=col,row=3).column].width = adjusted_width
                    col += 1

        # remove the blank worksheet that is automatically created with the spreadsheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb.get_sheet_by_name("Sheet"))

        wb.save(filename=outputLocation)


if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option("-s", "--schema", dest="schema_uri",
                      help="Base schema URI for the metadata")
    parser.add_option("-o", "--output", dest="output",
                      help="Output directory and file where to save the template spreadsheet", metavar="FILE")
    parser.add_option("-t", "--types", dest="schema_types",
                      help="Schema types to include in the spreadsheet")
    parser.add_option("-i", "--include", dest="include",
                      help="Schema modules to include in the spreadsheet")
    parser.add_option("-l", "--local", action="store_true", dest="local_files",
                      help="Run using local schemas")
    parser.add_option("-r", "--remote", action="store_false", dest="local_files",
                      help="Run using schemas from remote URL")
    parser.add_option("-u", "--user_friendly", action="store_true", dest="user_friendly",
                      help="Use user friendly field names")
    parser.add_option("-f", "--field_names", action="store_false", dest="user_friendly",
                      help="Use non-user friendly field names")

    (options, args) = parser.parse_args()

    # if not options.schema_uri:
    #     print ("You must supply a base schema URI for the metadata")
    #     exit(2)

    base_schema_path = options.schema_uri

    if options.local_files:

        core_schema_path = base_schema_path + '/core'
        type_schema_path = base_schema_path + '/type'
        module_schema_path = base_schema_path + '/module'

        core_schemas = [os.path.join(dirpath, f)
                        for dirpath, dirnames, files in os.walk(core_schema_path)
                        for f in files if f.endswith('.json')]

        schema_types = [os.path.join(dirpath, f)
                        for dirpath, dirnames, files in os.walk(type_schema_path)
                        for f in files if f.endswith('.json')]
        dependencies = [os.path.join(dirpath, f)
                          for dirpath, dirnames, files in os.walk(module_schema_path)
                          for f in files if f.endswith('.json')]

        ontologies = []

        for d in dependencies:
            if d.endswith('_ontology.json'):
                ontologies.append(d)
                dependencies.remove(d)


    else:
        schema_types = options.schema_types.split(",")
        dependencies = options.include.split(",")

        for index, dependency in enumerate(dependencies):
            dependencies[index] = options.schema_uri+dependency

    generator = SpreadsheetCreator()
    generator.generateSpreadsheet(base_schema_path, schema_types, dependencies, options.output, options.local_files, options.user_friendly)



# Example run:
# -s "https://schema.humancellatlas.org/"
# -t "type/biomaterial/donor_organism.json,type/process/sequencing/library_preparation_process.json"
# -i "module/biomaterial/homo_sapiens_specific.json,module/biomaterial/familial_relationship.json,module/process/sequencing/barcode.json"
# -o "/Users/dwelter/Development/HCA/metadata-schema/src/spreadsheet_test.xlsx"

# Full run (plain headers):
# -s "https://schema.humancellatlas.org/"
# -t "type/project/project.json,type/biomaterial/donor_organism.json,type/biomaterial/specimen_from_organism.json,type/biomaterial/cell_suspension.json,type/biomaterial/cell_line.json,type/biomaterial/organoid.json,type/process/biomaterial_collection/collection_process.json,type/process/biomaterial_collection/dissociation_process.json,type/process/biomaterial_collection/enrichment_process.json,type/process/sequencing/library_preparation_process.json,type/process/sequencing/sequencing_process.json,type/protocol/protocol.json,type/file/sequence_file.json"
# -i "module/project/contact.json,module/project/publication.json,module/biomaterial/cell_morphology.json,module/biomaterial/death.json,module/biomaterial/homo_sapiens_specific.json,module/biomaterial/medical_history.json,module/biomaterial/mus_musculus_specific.json,module/biomaterial/state_of_specimen.json,module/biomaterial/familial_relationship.json,module/process/sequencing/barcode.json,module/process/sequencing/smartseq2.json"
# -o "../examples/spreadsheets/v5/template/Programmatically_generated_v5_template_plainHeaders.xlsx"
# --field_names

# Full run (user-friendly headers):
# -s "https://schema.humancellatlas.org/"
# -t "type/project/project.json,type/biomaterial/donor_organism.json,type/biomaterial/specimen_from_organism.json,type/biomaterial/cell_suspension.json,type/biomaterial/cell_line.json,type/biomaterial/organoid.json,type/process/biomaterial_collection/collection_process.json,type/process/biomaterial_collection/dissociation_process.json,type/process/biomaterial_collection/enrichment_process.json,type/process/sequencing/library_preparation_process.json,type/process/sequencing/sequencing_process.json,type/protocol/protocol.json,type/file/sequence_file.json"
# -i "module/project/contact.json,module/project/publication.json,module/biomaterial/cell_morphology.json,module/biomaterial/death.json,module/biomaterial/homo_sapiens_specific.json,module/biomaterial/medical_history.json,module/biomaterial/mus_musculus_specific.json,module/biomaterial/state_of_specimen.json,module/biomaterial/familial_relationship.json,module/process/sequencing/barcode.json,module/process/sequencing/smartseq2.json"
# -o "../examples/spreadsheets/v5/template/Programmatically_generated_v5_template_spreadsheet.xlsx"
# --user_friendly

# Run using local files only
# -s "../json"
# -l
# -o "/Users/dwelter/Development/HCA/metadata-schema/src/spreadsheet_test.xlsx"
